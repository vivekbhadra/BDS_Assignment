#!/usr/bin/env python3
from openpyxl import load_workbook
import csv
import os

def convert_xlsx_to_csv(input_file, output_file, sheet_name=None, chunksize=100000):
    """
    Converts a large .xlsx file to .csv efficiently.
    Uses streaming with openpyxl to avoid high memory usage.
    
    :param input_file: Path to the input .xlsx file
    :param output_file: Path to the output .csv file
    :param sheet_name: Name of the sheet to convert (default: first sheet)
    :param chunksize: Number of rows to write at a time
    """
    
    # Load the workbook in read-only mode
    wb = load_workbook(filename=input_file, read_only=True)
    
    # Select the sheet
    sheet = wb[sheet_name] if sheet_name else wb.active

    # Open the CSV file for writing
    with open(output_file, mode="w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # Read the rows in chunks
        row_count = 0
        batch = []
        
        for row in sheet.iter_rows(values_only=True):
            batch.append(row)
            row_count += 1

            # Write in chunks
            if row_count % chunksize == 0:
                writer.writerows(batch)
                batch = []  # Reset batch

        # Write any remaining rows
        if batch:
            writer.writerows(batch)

    wb.close()
    print(f"Conversion completed: {input_file} → {output_file}")

# Example usage
convert_xlsx_to_csv("OnlineRetail.xlsx", "OnlineRetail.csv")
